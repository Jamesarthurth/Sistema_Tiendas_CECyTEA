import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Control de Tiendas Escolares", page_icon="📊", layout="wide")

TARIFAS_BASE = "tarifas_base.xlsx"
VERSION_REPORTE = "2.1"


def aplicar_fifo(pagos, cuotas_mes):
    resultado = {}
    for mes, cuota in cuotas_mes.items():
        resultado[mes] = {
            "Esperado": float(cuota) if pd.notna(cuota) else 0,
            "Pagado": 0,
            "Fecha": None,
            "Estado": "Pendiente",
        }

    for pago in pagos:
        monto = float(pago["monto"]) if pd.notna(pago["monto"]) else 0
        fecha = pago["fecha"]

        for mes in cuotas_mes.keys():
            pendiente = resultado[mes]["Esperado"] - resultado[mes]["Pagado"]
            if monto <= 0:
                break
            if pendiente <= 0:
                continue
            aplicado = min(monto, pendiente)
            resultado[mes]["Pagado"] += aplicado
            resultado[mes]["Fecha"] = fecha
            monto -= aplicado

    for mes in cuotas_mes.keys():
        esperado = resultado[mes]["Esperado"]
        pagado = resultado[mes]["Pagado"]
        if esperado == 0:
            resultado[mes]["Estado"] = "Sin cuota"
        elif pagado >= esperado:
            resultado[mes]["Estado"] = "Pagado"
        elif pagado > 0:
            resultado[mes]["Estado"] = "Parcial"
        else:
            resultado[mes]["Estado"] = "Pendiente"
    return resultado


def celda_monto_fecha(pagado, fecha, estado):
    if estado == "Sin cuota":
        return "SIN CUOTA"
    if estado == "Pendiente":
        return "🔴"
    monto_txt = f"${pagado:,.0f}"
    fecha_txt = "" if pd.isna(fecha) or fecha is None else pd.to_datetime(fecha).strftime("%d/%m")
    if estado == "Parcial":
        return f"{monto_txt}\n{fecha_txt}\n(P)"
    return f"{monto_txt}\n{fecha_txt}"


def detectar_meses_desde_cuotas(cuotas):
    """
    Detecta los meses del machote de cuotas.
    La estructura esperada es:
    columna 3 = primer mes cuota
    columna 4 = primer mes EE
    columna 5 = segundo mes cuota
    columna 6 = segundo mes EE
    y así sucesivamente.

    Ejemplo:
    Feb, Mar, Abr, May, Jun
    o
    Ago, Sep, Oct, Nov, Dic
    """
    meses_detectados = []

    for col in range(3, cuotas.shape[1], 2):
        valor = cuotas.iloc[0, col]

        if pd.isna(valor):
            continue

        mes = str(valor).strip()

        if mes and mes.lower() != "nan" and mes not in meses_detectados:
            meses_detectados.append(mes)

    if not meses_detectados:
        raise ValueError(
            "No se pudieron detectar meses en el archivo de cuotas. "
            "Revisa que los meses estén en la fila superior del machote."
        )

    return meses_detectados


def generar_reporte(df, cuotas):
    meses = detectar_meses_desde_cuotas(cuotas)

    columnas_necesarias = ["Nombre(s)", "  Fecha", "OTROS INGRESOS (ENERGIA ELEC)"]
    faltantes = [c for c in columnas_necesarias if c not in df.columns]
    if faltantes:
        raise ValueError(f"El archivo GLOBAL no tiene estas columnas necesarias: {faltantes}")

    columnas_cuota = [c for c in df.columns if "CUOTA RECUPERACION" in str(c)]
    if not columnas_cuota:
        raise ValueError("No se encontraron columnas de CUOTA RECUPERACION en el GLOBAL.")

    nombres = df["Nombre(s)"].astype(str)
    tiendas = df[
        nombres.str.contains("TIENDA", case=False, na=False)
        | nombres.str.contains("PABELLON DE ARTEAGA", case=False, na=False)
    ].copy()

    tiendas["  Fecha"] = pd.to_datetime(tiendas["  Fecha"], errors="coerce")
    tiendas["Pago_Cuota"] = tiendas[columnas_cuota].sum(axis=1, skipna=True)
    tiendas["Pago_EE"] = tiendas["OTROS INGRESOS (ENERGIA ELEC)"].fillna(0)

    def normalizar_tienda(nombre):
        nombre_original = str(nombre).strip().upper()

        # Excepciones específicas antes de limpiar
        if nombre_original == "CECYTEA TIENDA ESCOLAR":
            return "JESUS MARIA"

        if nombre_original == "EMS JESUS MARIA TIENDA ESCOLAR":
            return "EMS JESUS MARIA"

        # Limpieza general
        nombre_limpio = (
            nombre_original
            .replace("TIENDA ESCOLAR", "")
            .replace("CECYTEA", "")
            .replace("CECYT", "")
            .replace("EMS", "")
            .strip()
        )

        return nombre_limpio

    tiendas["Tienda"] = tiendas["Nombre(s)"].apply(normalizar_tienda)

    movimientos_tiendas = tiendas[(tiendas["Pago_Cuota"] > 0) | (tiendas["Pago_EE"] > 0)].copy()

    mapa_nombres = {
        "AGUASCALIENTES": "AGUASCALIENTES",
        "ASIENTOS": "ASIENTOS",
        "CAÑADA HONDA": "CAÑADA HONDA",
        "CECTY JESUS TERAN": "JESÚS TERÁN",
        "EL CHAYOTE": "EL CHAYOTE",
        "EL LLANO": "EL LLANO",
        "FERROCARRILES": "FERROCARRILES",
        "IGNACIO ZARAGOZA": "IGNACIO ZARAGOZA",
        "JESUS MARIA": "JESÚS MARÍA",
        "LAS FRAGUAS": "LAS FRAGUAS",
        "MIRADOR DE LAS CULTURAS": "MIRADOR DE LAS CULTURAS",
        "MORELOS": "MORELOS",
        "PABELLON DE HIDALGO": "PABELLON DE HIDALGO",
        "PAH": "PABELLON DE HIDALGO",
        "PABELLON DE ARTEAGA": 'PABELLON " NUEVO "',
        "RINCON DE ROMOS": "RINCÓN DE ROMOS",
        "SAN FCO DE LOS ROMO": "SAN FCO DE LOS ROMO",
        "SAN IGNACIO": "SAN IGNACIO",
        "SAN JOSE DE GRACIA": "SAN JOSÉ DE GRACIA",
        "EA CALVILLO": "CALVILLO",
        "EA VILLAMONTAÑA": 'VILLA MONTAÑA "NUEVO"',
        "EA": "EMS JESUS MARÍA",
        "VILLAMONTAÑA": 'VILLA MONTAÑA "NUEVO"',
        "CALVILLO": "CALVILLO",
        "EMS JESUS MARIA": "EMS JESUS MARÍA",
    }
    movimientos_tiendas["Plantel"] = movimientos_tiendas["Tienda"].map(mapa_nombres)
    sin_emparejar = movimientos_tiendas[movimientos_tiendas["Plantel"].isna()].copy()

    datos_cuotas = []
    for fila in range(2, len(cuotas), 2):
        plantel = cuotas.iloc[fila, 1]
        if pd.isna(plantel):
            continue
        registro = {"Plantel": plantel}

        for i, mes in enumerate(meses):
            col_cuota = 3 + (i * 2)
            col_ee = 4 + (i * 2)

            if col_ee >= cuotas.shape[1]:
                raise ValueError(
                    f"El machote de cuotas no tiene columnas suficientes para el mes {mes}."
                )

            registro[f"{mes}_Cuota"] = cuotas.iloc[fila, col_cuota]
            registro[f"{mes}_EE"] = cuotas.iloc[fila, col_ee]

        datos_cuotas.append(registro)
    cuotas_maestra = pd.DataFrame(datos_cuotas)

    reporte_filas = []
    for _, fila_cuota in cuotas_maestra.iterrows():
        plantel = fila_cuota["Plantel"]
        movimientos_plantel = movimientos_tiendas[movimientos_tiendas["Plantel"] == plantel].sort_values("  Fecha")
        cuotas_cuota = {mes: fila_cuota[f"{mes}_Cuota"] for mes in meses}
        cuotas_ee = {mes: fila_cuota[f"{mes}_EE"] for mes in meses}
        pagos_cuota = [{"fecha": row["  Fecha"], "monto": row["Pago_Cuota"]} for _, row in movimientos_plantel.iterrows()]
        pagos_ee = [{"fecha": row["  Fecha"], "monto": row["Pago_EE"]} for _, row in movimientos_plantel.iterrows()]
        resultado_cuota = aplicar_fifo(pagos_cuota, cuotas_cuota)
        resultado_ee = aplicar_fifo(pagos_ee, cuotas_ee)
        fila_reporte = {"Plantel": plantel}
        for mes in meses:
            fila_reporte[f"{mes}_Cuota_Estado"] = resultado_cuota[mes]["Estado"]
            fila_reporte[f"{mes}_Cuota_Fecha"] = resultado_cuota[mes]["Fecha"]
            fila_reporte[f"{mes}_Cuota_Pagado"] = resultado_cuota[mes]["Pagado"]
            fila_reporte[f"{mes}_Cuota_Esperado"] = resultado_cuota[mes]["Esperado"]
            fila_reporte[f"{mes}_EE_Estado"] = resultado_ee[mes]["Estado"]
            fila_reporte[f"{mes}_EE_Fecha"] = resultado_ee[mes]["Fecha"]
            fila_reporte[f"{mes}_EE_Pagado"] = resultado_ee[mes]["Pagado"]
            fila_reporte[f"{mes}_EE_Esperado"] = resultado_ee[mes]["Esperado"]
        reporte_filas.append(fila_reporte)
    reporte_fifo = pd.DataFrame(reporte_filas)

    filas_ejecutivo = []
    for _, row in reporte_fifo.iterrows():
        fila_cuota = {"Plantel": row["Plantel"], "Concepto": "CUOTA"}
        fila_ee = {"Plantel": row["Plantel"], "Concepto": "EE"}
        for mes in meses:
            fila_cuota[mes] = celda_monto_fecha(row[f"{mes}_Cuota_Pagado"], row[f"{mes}_Cuota_Fecha"], row[f"{mes}_Cuota_Estado"])
            fila_ee[mes] = celda_monto_fecha(row[f"{mes}_EE_Pagado"], row[f"{mes}_EE_Fecha"], row[f"{mes}_EE_Estado"])
        filas_ejecutivo.append(fila_cuota)
        filas_ejecutivo.append(fila_ee)
    reporte_ejecutivo = pd.DataFrame(filas_ejecutivo)

    filas_adeudos = []
    for _, row in reporte_fifo.iterrows():
        adeudo_cuota = adeudo_ee = 0
        meses_pendientes = []
        for mes in meses:
            falta_cuota = max(row[f"{mes}_Cuota_Esperado"] - row[f"{mes}_Cuota_Pagado"], 0)
            falta_ee = max(row[f"{mes}_EE_Esperado"] - row[f"{mes}_EE_Pagado"], 0)
            adeudo_cuota += falta_cuota
            adeudo_ee += falta_ee
            if falta_cuota > 0 or falta_ee > 0:
                meses_pendientes.append(mes)
        total_adeudo = adeudo_cuota + adeudo_ee
        estado_general = "Al corriente" if total_adeudo == 0 else ("Parcial" if len(meses_pendientes) <= 2 else "Con adeudo")
        filas_adeudos.append({
            "Plantel": row["Plantel"],
            "Adeudo Cuota": adeudo_cuota,
            "Adeudo EE": adeudo_ee,
            "Adeudo Total": total_adeudo,
            "Meses Pendientes": ", ".join(meses_pendientes),
            "Estado General": estado_general,
        })
    reporte_adeudos = pd.DataFrame(filas_adeudos).sort_values("Adeudo Total", ascending=False)

    fecha_mexico = datetime.now(ZoneInfo("America/Mexico_City")).strftime("%d/%m/%Y %H:%M")
    total_planteles = len(reporte_adeudos)
    planteles_al_corriente = (reporte_adeudos["Adeudo Total"] == 0).sum()
    planteles_con_adeudo = (reporte_adeudos["Adeudo Total"] > 0).sum()
    resumen = pd.DataFrame({
        "Indicador": [
            "Versión del reporte", "Fecha de generación", "Meses del reporte",
            "Total de planteles", "Planteles al corriente", "Planteles con adeudo",
            "Porcentaje al corriente", "Adeudo total", "Adeudo cuota", "Adeudo EE"
        ],
        "Valor": [
            VERSION_REPORTE, fecha_mexico, ", ".join(meses), total_planteles,
            planteles_al_corriente, planteles_con_adeudo,
            planteles_al_corriente / total_planteles if total_planteles else 0,
            reporte_adeudos["Adeudo Total"].sum(),
            reporte_adeudos["Adeudo Cuota"].sum(),
            reporte_adeudos["Adeudo EE"].sum(),
        ]
    })

    filas_detalle = []
    for _, row in reporte_fifo.iterrows():
        fila = {"Plantel": row["Plantel"]}
        for mes in meses:
            fila[f"{mes} Cuota Pagado"] = row[f"{mes}_Cuota_Pagado"]
            fila[f"{mes} Cuota Fecha"] = row[f"{mes}_Cuota_Fecha"]
            fila[f"{mes} Cuota Estado"] = row[f"{mes}_Cuota_Estado"]
            fila[f"{mes} EE Pagado"] = row[f"{mes}_EE_Pagado"]
            fila[f"{mes} EE Fecha"] = row[f"{mes}_EE_Fecha"]
            fila[f"{mes} EE Estado"] = row[f"{mes}_EE_Estado"]
        filas_detalle.append(fila)
    detalle_cobranza = pd.DataFrame(filas_detalle)

    salida_inicial = BytesIO()
    with pd.ExcelWriter(salida_inicial, engine="openpyxl") as writer:
        resumen.to_excel(writer, index=False, sheet_name="Resumen")
        reporte_ejecutivo.to_excel(writer, index=False, sheet_name="Reporte Ejecutivo")
        reporte_adeudos.to_excel(writer, index=False, sheet_name="Adeudos")
        detalle_cobranza.to_excel(writer, index=False, sheet_name="Detalle Cobranza")

    salida_inicial.seek(0)
    wb = load_workbook(salida_inicial)
    verde = PatternFill("solid", fgColor="C6EFCE")
    amarillo = PatternFill("solid", fgColor="FFEB9C")
    rojo = PatternFill("solid", fgColor="FFC7CE")
    gris = PatternFill("solid", fgColor="D9E1F2")
    azul = PatternFill("solid", fgColor="1F4E78")
    font_blanco = Font(color="FFFFFF", bold=True)
    font_normal = Font(color="000000")
    borde = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = azul
            cell.font = font_blanco
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borde
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = borde
                cell.font = font_normal
                valor = str(cell.value)
                if "🔴" in valor or valor == "Con adeudo":
                    cell.fill = rojo
                elif "(P)" in valor or valor == "Parcial":
                    cell.fill = amarillo
                elif "$" in valor or valor == "Al corriente":
                    cell.fill = verde
                elif "SIN CUOTA" in valor:
                    cell.fill = gris
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 25)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 25

    ws_ejecutivo = wb["Reporte Ejecutivo"]
    ws_ejecutivo.freeze_panes = "C2"
    for fila in range(2, ws_ejecutivo.max_row + 1):
        ws_ejecutivo.row_dimensions[fila].height = 40
    ws_ejecutivo.column_dimensions["A"].width = 24
    ws_ejecutivo.column_dimensions["B"].width = 12
    for col_idx in range(3, ws_ejecutivo.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws_ejecutivo.column_dimensions[col_letter].width = 14

    ws_adeudos = wb["Adeudos"]
    for row in ws_adeudos.iter_rows(min_row=2):
        for cell in row:
            if cell.column in [2, 3, 4]:
                cell.number_format = '$#,##0.00'

    ws_resumen = wb["Resumen"]
    for row in ws_resumen.iter_rows(min_row=2):
        indicador = row[0].value
        valor = row[1]
        if indicador in ["Adeudo total", "Adeudo cuota", "Adeudo EE"]:
            valor.number_format = '$#,##0.00'
        if indicador == "Porcentaje al corriente":
            valor.number_format = '0.00%'

    ws_detalle = wb["Detalle Cobranza"]
    for row in ws_detalle.iter_rows(min_row=2):
        for cell in row:
            encabezado = ws_detalle.cell(row=1, column=cell.column).value
            if "Pagado" in str(encabezado):
                cell.number_format = '$#,##0.00'
            if "Fecha" in str(encabezado):
                cell.number_format = 'dd/mm/yyyy'

    salida_final = BytesIO()
    wb.save(salida_final)
    salida_final.seek(0)

    return {
        "excel": salida_final,
        "resumen": resumen,
        "reporte_adeudos": reporte_adeudos,
        "sin_emparejar": sin_emparejar,
    }



# ============================================================
# INTERFAZ STREAMLIT - VERSIÓN PROFESIONAL
# ============================================================

import os

LOGO_PATH = "logo_cecytea.png"

st.markdown(
    """
    <style>
    :root {
        --cecytea-blue: #2B247C;
        --cecytea-lime: #CBE300;
        --soft-bg: #F7F8FC;
    }

    .main .block-container {
        padding-top: 1.5rem;
        padding-bottom: 2rem;
        max-width: 1200px;
    }

    .hero-card {
        background: linear-gradient(135deg, #ffffff 0%, #f6f7ff 100%);
        border: 1px solid #dddff5;
        border-radius: 18px;
        padding: 28px 30px;
        box-shadow: 0 6px 18px rgba(43,36,124,0.08);
        margin-bottom: 24px;
    }

    .hero-title {
        color: #2B247C;
        font-size: 34px;
        font-weight: 800;
        margin-bottom: 6px;
        line-height: 1.15;
    }

    .hero-subtitle {
        color: #4B4B63;
        font-size: 17px;
        margin-bottom: 0;
    }

    .section-title {
        border-left: 6px solid #CBE300;
        padding-left: 12px;
        color: #2B247C;
        font-size: 23px;
        font-weight: 800;
        margin-top: 20px;
        margin-bottom: 10px;
    }

    .info-card {
        background: #ffffff;
        border: 1px solid #e5e7f5;
        border-radius: 14px;
        padding: 18px 20px;
        box-shadow: 0 4px 12px rgba(31,78,121,0.06);
        height: 100%;
    }

    .step-number {
        background: #2B247C;
        color: white;
        border-radius: 50%;
        padding: 6px 12px;
        font-weight: 700;
        margin-right: 8px;
    }

    .footer {
        margin-top: 35px;
        padding: 18px;
        border-top: 5px solid #CBE300;
        background: #2B247C;
        color: white;
        text-align: center;
        border-radius: 12px;
        font-size: 14px;
    }

    div.stButton > button:first-child {
        background-color: #2B247C;
        color: white;
        border-radius: 10px;
        border: none;
        padding: 0.6rem 1.2rem;
        font-weight: 700;
    }

    div.stButton > button:first-child:hover {
        background-color: #1f1a63;
        color: white;
        border: none;
    }

    div.stDownloadButton > button:first-child {
        background-color: #2B247C;
        color: white;
        border-radius: 10px;
        border: none;
        padding: 0.7rem 1.2rem;
        font-weight: 700;
    }

    div.stDownloadButton > button:first-child:hover {
        background-color: #1f1a63;
        color: white;
        border: none;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# -------------------------
# SIDEBAR
# -------------------------

with st.sidebar:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, use_container_width=True)
    else:
        st.markdown("### CECyTEA")

    st.markdown("---")
    st.markdown("### ⚙️ Configuración")
    st.caption("Sistema de Control de Tiendas Escolares")

    st.markdown("#### 📁 Cuotas base activas")
    st.info(f"Archivo activo: `{TARIFAS_BASE}`")

    # Botón para descargar machote de cuotas
    try:
        with open(TARIFAS_BASE, "rb") as archivo_machote:
            st.download_button(
                label="📥 Descargar machote de cuotas",
                data=archivo_machote,
                file_name="Machote_Cuotas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except FileNotFoundError:
        st.error("No se encontró el archivo de cuotas base.")

    actualizar_cuotas = st.file_uploader(
        "Actualizar archivo de cuotas base",
        type=["xlsx"],
        key="cuotas",
        help="Sube un nuevo Excel de cuotas únicamente cuando cambie el periodo o las tarifas."
    )

    if actualizar_cuotas is not None:
        st.warning("Esta acción reemplazará las cuotas base actuales.")
        if st.button("Guardar nuevas cuotas base"):
            with open(TARIFAS_BASE, "wb") as f:
                f.write(actualizar_cuotas.getbuffer())
            st.success("Cuotas base actualizadas correctamente.")

    st.markdown("---")
    st.markdown("#### ℹ️ Información")
    st.write(f"**Versión:** {VERSION_REPORTE}")
    st.write("**Método:** FIFO")
    st.caption("Los pagos se aplican primero al mes adeudado más antiguo.")

# -------------------------
# ENCABEZADO PRINCIPAL
# -------------------------

if os.path.exists(LOGO_PATH):
    col_logo, col_text = st.columns([1, 4])
    with col_logo:
        st.image(LOGO_PATH, use_container_width=True)
    with col_text:
        st.markdown(
            """
            <div class="hero-card">
                <div class="hero-title">Sistema de Control de Tiendas Escolares</div>
                <p class="hero-subtitle">
                    Generación automática de reportes ejecutivos de pagos, adeudos y detalle de cobranza.
                </p>
            </div>
            """,
            unsafe_allow_html=True
        )
else:
    st.markdown(
        """
        <div class="hero-card">
            <div class="hero-title">📊 Sistema de Control de Tiendas Escolares</div>
            <p class="hero-subtitle">
                Generación automática de reportes ejecutivos de pagos, adeudos y detalle de cobranza.
            </p>
        </div>
        """,
        unsafe_allow_html=True
    )

st.markdown('<div class="section-title">¿Cómo usar el sistema?</div>', unsafe_allow_html=True)

col_a, col_b, col_c = st.columns(3)
with col_a:
    st.markdown(
        """
        <div class="info-card">
            <b><span class="step-number">1</span>Sube GLOBAL</b><br><br>
            Carga el archivo GLOBAL en formato .xlsx. Debe contener la hoja <b>2024</b>.
        </div>
        """,
        unsafe_allow_html=True
    )
with col_b:
    st.markdown(
        """
        <div class="info-card">
            <b><span class="step-number">2</span>Genera reporte</b><br><br>
            El sistema procesa pagos, aplica FIFO y calcula adeudos automáticamente.
        </div>
        """,
        unsafe_allow_html=True
    )
with col_c:
    st.markdown(
        """
        <div class="info-card">
            <b><span class="step-number">3</span>Descarga Excel</b><br><br>
            Obtén un archivo con Resumen, Reporte Ejecutivo, Adeudos y Detalle de Cobranza.
        </div>
        """,
        unsafe_allow_html=True
    )

st.markdown('<div class="section-title">1. Subir archivo GLOBAL</div>', unsafe_allow_html=True)

archivo_global = st.file_uploader(
    "Selecciona el archivo GLOBAL",
    type=["xlsx"],
    key="global",
    help="Archivo de movimientos. Debe contener la hoja llamada 2024."
)

if archivo_global is not None:
    try:
        df = pd.read_excel(archivo_global, sheet_name="2024", header=1)
        cuotas = pd.read_excel(TARIFAS_BASE, sheet_name=0, header=None)
        meses_detectados = detectar_meses_desde_cuotas(cuotas)

        st.success("Archivo GLOBAL y cuotas base leídos correctamente.")
        st.info("Meses detectados en el machote de cuotas: " + ", ".join(meses_detectados))

        st.markdown('<div class="section-title">Estado del sistema</div>', unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        col1.metric("Registros GLOBAL", f"{df.shape[0]:,}")
        col2.metric("Columnas GLOBAL", df.shape[1])
        col3.metric("Meses detectados", len(meses_detectados))

        st.markdown('<div class="section-title">2. Generar reporte</div>', unsafe_allow_html=True)

        if st.button("🚀 Generar reporte ejecutivo"):
            with st.spinner("Generando reporte, por favor espera..."):
                resultado = generar_reporte(df, cuotas)

            if len(resultado["sin_emparejar"]) > 0:
                st.warning("Hay tiendas sin emparejar. Revisa la tabla inferior.")
                st.dataframe(resultado["sin_emparejar"], use_container_width=True)

            st.success("Reporte generado correctamente.")

            st.markdown('<div class="section-title">Resumen del reporte</div>', unsafe_allow_html=True)

            resumen_df = resultado["resumen"].copy()
            st.dataframe(resumen_df, use_container_width=True, hide_index=True)

            try:
                adeudos = resultado["reporte_adeudos"]
                total_planteles = len(adeudos)
                al_corriente = (adeudos["Adeudo Total"] == 0).sum()
                con_adeudo = (adeudos["Adeudo Total"] > 0).sum()
                adeudo_total = adeudos["Adeudo Total"].sum()

                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Planteles", total_planteles)
                m2.metric("Al corriente", al_corriente)
                m3.metric("Con adeudo", con_adeudo)
                m4.metric("Adeudo total", f"${adeudo_total:,.2f}")
            except Exception:
                pass

            st.markdown('<div class="section-title">Top 10 adeudos</div>', unsafe_allow_html=True)
            st.dataframe(resultado["reporte_adeudos"].head(10), use_container_width=True, hide_index=True)

            st.markdown('<div class="section-title">3. Descargar reporte</div>', unsafe_allow_html=True)
            st.download_button(
                label="📥 Descargar reporte Excel",
                data=resultado["excel"],
                file_name="Reporte_Ejecutivo_Tiendas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"Ocurrió un error: {e}")
else:
    st.info("Sube un archivo GLOBAL para comenzar.")

st.markdown(
    """
    <div class="footer">
        <b>CECyTEA | Sistema de Control de Tiendas Escolares</b><br>
        Versión 2.1 · Meses dinámicos desde machote · Método FIFO
    </div>
    """,
    unsafe_allow_html=True
)
